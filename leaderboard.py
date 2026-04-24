"""Leaderboard display and management."""
from tabulate import tabulate
from storage import load_scores, load_match_history, load_scoring_sheet, load_master_scoresheet, get_cached_scorecard


def _short_match_name_from_description(description: str) -> str:
    if not description:
        return ""

    label = description.strip()
    if label.lower().startswith("cricket scorecard - "):
        label = label[len("cricket scorecard - "):].strip()

    for sep in [" – ", " - ", " | "]:
        if sep in label:
            label = label.split(sep, 1)[0].strip()
            break

    if "," in label and " vs " in label:
        label = label.split(",", 1)[0].strip()

    return label


def _short_match_name_from_scorecard(scorecard_data: dict, match_id: int = 0) -> str:
    if not scorecard_data:
        return f"Match {match_id}"

    header = scorecard_data.get("matchHeader") or scorecard_data.get("match_header") or {}
    if isinstance(header, dict):
        t1 = header.get("team1", {}).get("shortName") or header.get("team1", {}).get("shortname")
        t2 = header.get("team2", {}).get("shortName") or header.get("team2", {}).get("shortname")
        if t1 and t2:
            return f"{t1} vs {t2}"

    appindex = scorecard_data.get("appindex", {}) or scorecard_data.get("appIndex", {})
    seo = appindex.get("seotitle") or appindex.get("seoTitle") or ""
    if seo:
        part = seo.split("|", 1)[0].strip()
        if part.lower().startswith("cricket scorecard - "):
            part = part[len("cricket scorecard - "):].strip()
        return _short_match_name_from_description(part)

    if scorecard_data.get("matchDesc"):
        return _short_match_name_from_description(scorecard_data["matchDesc"])

    if scorecard_data.get("description"):
        return _short_match_name_from_description(scorecard_data["description"])

    return f"Match {match_id}"


def _match_label_for_master_entry(match: dict) -> str:
    description = match.get("description", "")
    short = _short_match_name_from_description(description)

    if not short or short.lower().startswith("match "):
        cached = get_cached_scorecard(match.get("match_id"))
        if cached:
            short = _short_match_name_from_scorecard(cached, match.get("match_id"))

    if not short:
        return f"Match {match.get('match_id')}"
    return short


def show_leaderboard():
    """Print the cumulative leaderboard sorted by total points."""
    scores = load_scores()
    history = load_match_history()
    matches_played = len(history)

    # Sort teams by total points descending
    sorted_teams = sorted(scores.items(), key=lambda x: x[1], reverse=True)

    table = []
    for rank, (team, points) in enumerate(sorted_teams, 1):
        table.append([rank, team, round(points, 2)])

    print(f"\n{'='*50}")
    print(f"  IPL FANTASY PREMIER LEAGUE 2026 – LEADERBOARD")
    print(f"  Matches processed: {matches_played}")
    print(f"{'='*50}")
    print(tabulate(table, headers=["Rank", "Team Owner", "Total Points"],
                   tablefmt="fancy_grid", numalign="right"))
    print()


def show_match_history():
    """Print all processed matches with per-team scores."""
    history = load_match_history()
    if not history:
        print("\nNo matches processed yet.\n")
        return

    for i, match in enumerate(history, 1):
        print(f"\n── Match {i}: {match['description']} (ID: {match['match_id']}) ──")
        sorted_scores = sorted(match["team_scores"].items(),
                                key=lambda x: x[1], reverse=True)
        table = [[rank, team, round(pts, 2)]
                 for rank, (team, pts) in enumerate(sorted_scores, 1)]
        print(tabulate(table, headers=["#", "Team", "Points"],
                       tablefmt="simple", numalign="right"))
    print()


def show_match_detail(match_results: dict, description: str = ""):
    """Print detailed per-player breakdown for a single match."""
    if description:
        print(f"\n── {description} ──")

    sorted_teams = sorted(match_results.items(),
                          key=lambda x: x[1]["total"], reverse=True)

    for rank, (owner, data) in enumerate(sorted_teams, 1):
        print(f"\n  {rank}. {owner}: {data['total']} pts")
        if data["players"]:
            player_table = []
            for pname, pdata in sorted(data["players"].items(),
                                        key=lambda x: x[1]["total"],
                                        reverse=True):
                mult = ""
                if pdata["multiplier"] == 1.5:
                    mult = " (C)"
                elif pdata["multiplier"] == 1.25:
                    mult = " (VC)"
                player_table.append([
                    f"{pname}{mult}",
                    pdata["batting"], pdata["bowling"],
                    pdata["fielding"], pdata["generic"],
                    pdata["base_total"], pdata["total"],
                ])
            print(tabulate(
                player_table,
                headers=["Player", "Bat", "Bowl", "Field", "Gen",
                         "Base", "Total"],
                tablefmt="simple", numalign="right", floatfmt=".1f",
            ))
    print()


def show_scoring_sheet():
    """Print the full scoring sheet: match-by-match with team totals and cumulative progression."""
    sheet = load_scoring_sheet()
    matches = sheet.get("matches", [])
    if not matches:
        print("\nNo matches in scoring sheet yet.\n")
        return

    # Collect all team owners
    owners = list(matches[0]["team_results"].keys()) if matches else []

    # ── Summary table: how each team's score progresses match by match ──
    print(f"\n{'='*70}")
    print(f"  SCORING SHEET – CUMULATIVE PROGRESSION")
    print(f"{'='*70}")

    header = ["Team"] + [f"M{i+1}" for i in range(len(matches))] + ["Total"]
    rows = []
    for owner in owners:
        row = [owner]
        for m in matches:
            tr = m["team_results"].get(owner, {})
            row.append(round(tr.get("match_points", 0), 1))
        cum = matches[-1]["team_results"].get(owner, {}).get("cumulative_points", 0)
        row.append(round(cum, 1))
        rows.append(row)

    # Sort by cumulative (last column) descending
    rows.sort(key=lambda r: r[-1], reverse=True)
    print(tabulate(rows, headers=header, tablefmt="fancy_grid",
                   numalign="right", floatfmt=".1f"))

    # ── Match legend ──
    print("\nMatch key:")
    for i, m in enumerate(matches):
        print(f"  M{i+1}: {m['description']} (ID: {m['match_id']})")

    # ── Per-match detail ──
    for i, m in enumerate(matches):
        print(f"\n{'─'*60}")
        print(f"  M{i+1}: {m['description']}")
        print(f"{'─'*60}")

        sorted_teams = sorted(
            m["team_results"].items(),
            key=lambda x: x[1]["match_points"], reverse=True
        )
        for rank, (owner, tr) in enumerate(sorted_teams, 1):
            cum = round(tr.get("cumulative_points", 0), 1)
            mp = round(tr.get("match_points", 0), 1)
            print(f"\n  {rank}. {owner}: {mp} pts (cumulative: {cum})")

            players = tr.get("players", {})
            if players:
                ptable = []
                for pname, pd in sorted(players.items(),
                                          key=lambda x: x[1]["total"],
                                          reverse=True):
                    mult = ""
                    if pd["multiplier"] == 1.5:
                        mult = " (C)"
                    elif pd["multiplier"] == 1.25:
                        mult = " (VC)"
                    ptable.append([
                        f"{pname}{mult}",
                        pd["batting"], pd["bowling"],
                        pd["fielding"], pd["generic"],
                        pd["base_total"], pd["total"],
                    ])
                print(tabulate(
                    ptable,
                    headers=["Player", "Bat", "Bowl", "Field", "Gen",
                             "Base", "Total"],
                    tablefmt="simple", numalign="right", floatfmt=".1f",
                ))
    print()


# ══════════════════════════════════════════════════════════════════════════════
# MASTER SCORESHEET – cumulative player-by-player view across all matches
# ══════════════════════════════════════════════════════════════════════════════

def show_master_scoresheet():
    """Print the master scoresheet: per-team, per-player totals across all matches."""
    master = load_master_scoresheet()
    match_list = master.get("match_list", [])
    teams = master.get("teams", {})

    if not match_list or not teams:
        print("\nNo data in master scoresheet yet. Process a match first.\n")
        return

    # ── Overall leaderboard ──
    print(f"\n{'='*80}")
    print(f"  MASTER SCORESHEET – IPL FANTASY PREMIER LEAGUE 2026")
    print(f"  Matches: {len(match_list)}")
    print(f"{'='*80}")

    # Team standings summary
    sorted_teams = sorted(teams.items(),
                          key=lambda x: x[1]["cumulative_total"], reverse=True)

    # Build header with match columns
    match_headers = []
    for i, m in enumerate(match_list):
        match_headers.append(f"M{i+1}")

    header = ["#", "Team"] + match_headers + ["Total"]
    rows = []
    for rank, (owner, tdata) in enumerate(sorted_teams, 1):
        row = [rank, owner]
        for m in match_list:
            mid = str(m["match_id"])
            match_total = sum(
                ps["match_scores"].get(mid, {}).get("total", 0)
                for ps in tdata["players"].values()
            )
            row.append(round(match_total, 1))
        row.append(round(tdata["cumulative_total"], 1))
        rows.append(row)

    print(tabulate(rows, headers=header, tablefmt="fancy_grid",
                   numalign="right", floatfmt=".1f"))

    # Match legend
    print("\n  Match key:")
    for i, m in enumerate(match_list):
        print(f"    M{i+1}: {m['description']} (ID: {m['match_id']})")

    # ── Per-team player breakdown ──
    for rank, (owner, tdata) in enumerate(sorted_teams, 1):
        print(f"\n{'━'*80}")
        print(f"  {rank}. {owner} — TOTAL: {tdata['cumulative_total']} pts")
        print(f"{'━'*80}")

        # Sort players: active ones first (by cumulative total), then inactive
        players = list(tdata["players"].items())
        played_players = [(n, p) for n, p in players
                          if p["cumulative"]["matches_played"] > 0 or p["cumulative"]["total"] != 0]
        unplayed = [(n, p) for n, p in players
                    if p["cumulative"]["matches_played"] == 0 and p["cumulative"]["total"] == 0]
        played_players.sort(key=lambda x: x[1]["cumulative"]["total"], reverse=True)
        unplayed.sort(key=lambda x: x[0])

        header = ["Player", "Role"] + match_headers + ["Bat", "Bowl", "Fld", "Gen", "Total", "MP"]
        rows = []
        for pname, pdata in played_players:
            tag = ""
            if pdata.get("is_captain"):
                tag = " (C)"
            elif pdata.get("is_vice_captain"):
                tag = " (VC)"
            row = [f"{pname}{tag}", pdata["role"]]
            for m in match_list:
                mid = str(m["match_id"])
                ms = pdata["match_scores"].get(mid)
                if ms and (ms.get("played") or ms.get("total", 0) != 0):
                    row.append(round(ms["total"], 1))
                else:
                    row.append("—")
            cum = pdata["cumulative"]
            row.extend([
                round(cum["batting"], 1),
                round(cum["bowling"], 1),
                round(cum["fielding"], 1),
                round(cum["generic"], 1),
                round(cum["total"], 1),
                cum["matches_played"],
            ])
            rows.append(row)

        for pname, pdata in unplayed:
            tag = ""
            if pdata.get("is_captain"):
                tag = " (C)"
            elif pdata.get("is_vice_captain"):
                tag = " (VC)"
            row = [f"{pname}{tag}", pdata["role"]]
            for _m in match_list:
                row.append("—")
            row.extend(["—", "—", "—", "—", "—", 0])
            rows.append(row)

        print(tabulate(rows, headers=header, tablefmt="simple",
                       numalign="right", floatfmt=".1f",
                       colalign=("left", "left") + ("right",) * (len(header) - 2)))

    print()


# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY VIEW – all teams with all players and their points
# ══════════════════════════════════════════════════════════════════════════════

def show_summary(match_results: dict, description: str = ""):
    """Print summary: every team with ALL players (played or not) and points."""
    if description:
        print(f"\n{'='*70}")
        print(f"  MATCH SUMMARY – {description}")
        print(f"{'='*70}")

    sorted_teams = sorted(match_results.items(),
                          key=lambda x: x[1]["total"], reverse=True)

    for rank, (owner, data) in enumerate(sorted_teams, 1):
        print(f"\n┌─ {rank}. {owner}: {data['total']} pts")
        print(f"├{'─'*68}")

        players = data.get("players", {})
        played = [(n, p) for n, p in players.items() if p.get("played")]
        not_played = [(n, p) for n, p in players.items() if not p.get("played")]
        played.sort(key=lambda x: x[1]["total"], reverse=True)
        not_played.sort(key=lambda x: x[0])

        rows = []
        for pname, pd in played:
            tag = _player_tag(pd)
            rows.append([
                f"  {pname}{tag}",
                pd["role"],
                pd["batting"], pd["bowling"], pd["fielding"], pd["generic"],
                pd["total"],
            ])
        for pname, pd in not_played:
            rows.append([
                f"  {pname}", pd["role"],
                "-", "-", "-", "-", "-",
            ])

        print(tabulate(
            rows,
            headers=["  Player", "Role", "Bat", "Bowl", "Field", "Gen", "Total"],
            tablefmt="simple", numalign="right", floatfmt=".1f",
            colalign=("left", "left", "right", "right", "right", "right", "right"),
        ))
        print(f"└{'─'*68}")
    print()


# ══════════════════════════════════════════════════════════════════════════════
# DETAILED VIEW – granular point breakdown per player per team
# ══════════════════════════════════════════════════════════════════════════════

def show_detailed(match_results: dict, description: str = ""):
    """Print detailed point breakdown for each player in each team."""
    if description:
        print(f"\n{'='*70}")
        print(f"  DETAILED SCORECARD – {description}")
        print(f"{'='*70}")

    sorted_teams = sorted(match_results.items(),
                          key=lambda x: x[1]["total"], reverse=True)

    for rank, (owner, data) in enumerate(sorted_teams, 1):
        print(f"\n{'━'*70}")
        print(f"  {rank}. {owner} — TOTAL: {data['total']} pts")
        print(f"{'━'*70}")

        players = data.get("players", {})
        played = [(n, p) for n, p in players.items() if p.get("played")]
        played.sort(key=lambda x: x[1]["total"], reverse=True)

        if not played:
            print("  No players featured in this match.")
            continue

        for pname, pd in played:
            tag = _player_tag(pd)
            raw = pd.get("raw", {})

            print(f"\n  ▸ {pname}{tag}  ({pd['role']})")
            print(f"    {'─'*50}")

            # Batting breakdown
            if "bat" in raw:
                b = raw["bat"]
                not_out = "*" if b["is_not_out"] else ""
                print(f"    BATTING:  {b['runs']}{not_out} ({b['balls']}b)  "
                      f"{b['fours']}×4  {b['sixes']}×6  SR {b['sr']}")

                lines = []
                lines.append(f"      Runs: {b['runs']}×1 = {b['runs']}")
                if b["fours"]:
                    lines.append(f"      4s bonus: {b['fours']}×1 = {b['fours']}")
                if b["sixes"]:
                    lines.append(f"      6s bonus: {b['sixes']}×2 = {b['sixes']*2}")

                r = b["runs"]
                if r >= 150:   lines.append(f"      150+ milestone: +30")
                elif r >= 125: lines.append(f"      125+ milestone: +25")
                elif r >= 100: lines.append(f"      100+ milestone: +20")
                elif r >= 75:  lines.append(f"      75+ milestone: +15")
                elif r >= 50:  lines.append(f"      50+ milestone: +10")
                elif r >= 30:  lines.append(f"      30+ milestone: +5")

                if r == 0 and not b["is_not_out"] and b["balls"] > 0:
                    if pd["role"] in ("Batsman", "Allrounder", "Wicketkeeper"):
                        lines.append(f"      Duck penalty: -5")

                if b["balls"] >= 10:
                    sr = b["sr"]
                    if sr > 200:    lines.append(f"      SR {sr} (>200): +12")
                    elif sr > 150:  lines.append(f"      SR {sr} (>150): +8")
                    elif sr >= 120: lines.append(f"      SR {sr} (120-150): +4")
                    elif sr >= 80:  lines.append(f"      SR {sr} (80-120): +0")
                    else:           lines.append(f"      SR {sr} (<80): -4")

                for l in lines:
                    print(l)
                print(f"      → Batting total: {pd['batting']}")

            # Bowling breakdown
            if "bowl" in raw:
                bw = raw["bowl"]
                print(f"    BOWLING:  {bw['overs']}ov  {bw['maidens']}m  "
                      f"{bw['runs']}r  {bw['wickets']}w  Econ {bw['econ']}")

                lines = []
                if bw["wickets"]:
                    lines.append(f"      Wickets: {bw['wickets']}×25 = {bw['wickets']*25}")
                if bw["bowled_lbw"]:
                    lines.append(f"      B/LBW/HW bonus: {bw['bowled_lbw']}×10 = {bw['bowled_lbw']*10}")

                w = bw["wickets"]
                if w >= 5:   lines.append(f"      5W haul: +32")
                elif w >= 4: lines.append(f"      4W haul: +24")
                elif w >= 3: lines.append(f"      3W haul: +16")
                elif w >= 2: lines.append(f"      2W haul: +8")

                if bw["maidens"]:
                    lines.append(f"      Maidens: {bw['maidens']}×8 = {bw['maidens']*8}")
                if bw["dot_balls"]:
                    lines.append(f"      Dots: {bw['dot_balls']}×0.5 = {bw['dot_balls']*0.5}")

                if bw["overs"] >= 1:
                    e = bw["econ"]
                    if e >= 12:    lines.append(f"      Econ {e} (≥12): -8")
                    elif e >= 9:   lines.append(f"      Econ {e} (9-12): -4")
                    elif e > 7.5:  lines.append(f"      Econ {e} (7.5-9): +0")
                    elif e > 6:    lines.append(f"      Econ {e} (6-7.5): +4")
                    elif e <= 5:   lines.append(f"      Econ {e} (≤5): +16")
                    else:          lines.append(f"      Econ {e} (5-6): +8")

                for l in lines:
                    print(l)
                print(f"      → Bowling total: {pd['bowling']}")

            # Fielding breakdown
            if "field" in raw:
                fl = raw["field"]
                parts = []
                if fl["catches"]:    parts.append(f"{fl['catches']}c")
                if fl["stumpings"]:  parts.append(f"{fl['stumpings']}st")
                if fl["run_out_solo"]:  parts.append(f"{fl['run_out_solo']} solo r/o")
                if fl["run_out_shared"]: parts.append(f"{fl['run_out_shared']} shared r/o")
                print(f"    FIELDING: {', '.join(parts)}")

                lines = []
                if fl["catches"]:
                    lines.append(f"      Catches: {fl['catches']}×10 = {fl['catches']*10}")
                if fl["stumpings"]:
                    lines.append(f"      Stumpings: {fl['stumpings']}×10 = {fl['stumpings']*10}")
                if fl["run_out_solo"]:
                    lines.append(f"      Solo run-out: {fl['run_out_solo']}×20 = {fl['run_out_solo']*20}")
                if fl["run_out_shared"]:
                    lines.append(f"      Shared run-out: {fl['run_out_shared']}×10 = {fl['run_out_shared']*10}")
                for l in lines:
                    print(l)
                print(f"      → Fielding total: {pd['fielding']}")

            # Generic
            gen_parts = []
            if pd.get("played"): gen_parts.append("Playing XI +4")
            if pd["generic"] >= 34: gen_parts.append("MoM +30")
            if gen_parts:
                print(f"    GENERIC:  {', '.join(gen_parts)}  → {pd['generic']}")

            # Total
            print(f"    {'─'*50}")
            base = pd["base_total"]
            mult = pd["multiplier"]
            if mult != 1.0:
                label = "Captain ×1.5" if mult == 1.5 else "Vice-Captain ×1.25"
                print(f"    BASE: {base}  ×  {label}  =  {pd['total']}")
            else:
                print(f"    TOTAL: {pd['total']}")

    print()


def _player_tag(pd: dict) -> str:
    """Return (C), (VC) or empty string."""
    if pd.get("multiplier") == 1.5:
        return " (C)"
    elif pd.get("multiplier") == 1.25:
        return " (VC)"
    return ""


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT – Summary and Detailed views
# ══════════════════════════════════════════════════════════════════════════════

def export_summary_to_excel(match_results: dict, filename: str = "summary.xlsx"):
    """Export summary view to Excel: all teams with all players."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    
    # Title
    ws['A1'] = "IPL Fantasy Premier League 2026 – Match Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    row = 3
    sorted_teams = sorted(match_results.items(), key=lambda x: x[1]["total"], reverse=True)
    
    for rank, (owner, data) in enumerate(sorted_teams, 1):
        # Team header
        ws[f'A{row}'] = f"{rank}. {owner}"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        # Column headers
        headers = ["Player", "Role", "Batting", "Bowling", "Fielding", "Generic", "Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        row += 1
        
        # Players
        players = data.get("players", {})
        played = [(n, p) for n, p in players.items() if p.get("played")]
        not_played = [(n, p) for n, p in players.items() if not p.get("played")]
        played.sort(key=lambda x: x[1]["total"], reverse=True)
        not_played.sort(key=lambda x: x[0])
        
        for pname, pd in played + not_played:
            tag = _player_tag(pd)
            ws.cell(row=row, column=1, value=f"{pname}{tag}").border = border
            ws.cell(row=row, column=2, value=pd["role"]).border = border
            
            if pd.get("played"):
                ws.cell(row=row, column=3, value=pd["batting"]).border = border
                ws.cell(row=row, column=4, value=pd["bowling"]).border = border
                ws.cell(row=row, column=5, value=pd["fielding"]).border = border
                ws.cell(row=row, column=6, value=pd["generic"]).border = border
                ws.cell(row=row, column=7, value=pd["total"]).border = border
                
                for col in [3, 4, 5, 6, 7]:
                    ws.cell(row=row, column=col).alignment = right_align
                    ws.cell(row=row, column=col).number_format = '0.0'
            else:
                for col in range(3, 8):
                    ws.cell(row=row, column=col, value="—").border = border
                    ws.cell(row=row, column=col).alignment = center_align
            
            ws.cell(row=row, column=1).alignment = left_align
            ws.cell(row=row, column=2).alignment = left_align
            row += 1

        # Totals row
        total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        total_font = Font(bold=True, size=10)
        ws.cell(row=row, column=1, value="TOTAL").border = border
        ws.cell(row=row, column=1).font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=2, value="").border = border
        ws.cell(row=row, column=2).fill = total_fill
        bat_sum = sum(pd["batting"] for _, pd in played)
        bowl_sum = sum(pd["bowling"] for _, pd in played)
        field_sum = sum(pd["fielding"] for _, pd in played)
        gen_sum = sum(pd["generic"] for _, pd in played)
        for ci, val in enumerate([bat_sum, bowl_sum, field_sum, gen_sum, data["total"]], 3):
            cell = ws.cell(row=row, column=ci, value=round(val, 1))
            cell.number_format = '0.0'
            cell.alignment = right_align
            cell.border = border
            cell.font = total_font
            cell.fill = total_fill
        row += 1

        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    
    wb.save(filename)
    print(f"✓ Summary exported to {filename}")


def export_detailed_to_excel(match_results: dict, filename: str = "detailed.xlsx"):
    """Export detailed view to Excel: one sheet per team with granular breakdowns."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    
    sorted_teams = sorted(match_results.items(), key=lambda x: x[1]["total"], reverse=True)
    
    for rank, (owner, data) in enumerate(sorted_teams, 1):
        ws = wb.create_sheet(title=owner[:31])  # Excel sheet name max 31 chars
        
        # Title
        ws['A1'] = f"{owner} – Total: {data['total']} pts"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:H1')
        
        row = 3
        
        # Column headers
        headers = ["Player", "Role", "Batting", "Bowling", "Fielding", "Generic", "Base", "Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        row += 1
        
        # Players
        players = data.get("players", {})
        played = [(n, p) for n, p in players.items() if p.get("played")]
        played.sort(key=lambda x: x[1]["total"], reverse=True)
        
        if not played:
            ws.cell(row=row, column=1, value="No players featured in this match").border = border
        else:
            for pname, pd in played:
                tag = _player_tag(pd)
                ws.cell(row=row, column=1, value=f"{pname}{tag}").border = border
                ws.cell(row=row, column=2, value=pd["role"]).border = border
                ws.cell(row=row, column=3, value=pd["batting"]).number_format = '0.0'
                ws.cell(row=row, column=3).border = border
                ws.cell(row=row, column=4, value=pd["bowling"]).number_format = '0.0'
                ws.cell(row=row, column=4).border = border
                ws.cell(row=row, column=5, value=pd["fielding"]).number_format = '0.0'
                ws.cell(row=row, column=5).border = border
                ws.cell(row=row, column=6, value=pd["generic"]).number_format = '0.0'
                ws.cell(row=row, column=6).border = border
                ws.cell(row=row, column=7, value=pd["base_total"]).number_format = '0.0'
                ws.cell(row=row, column=7).border = border
                ws.cell(row=row, column=8, value=pd["total"]).number_format = '0.0'
                ws.cell(row=row, column=8).border = border
                
                ws.cell(row=row, column=1).alignment = left_align
                ws.cell(row=row, column=2).alignment = left_align
                for col in range(3, 9):
                    ws.cell(row=row, column=col).alignment = right_align
                
                row += 1

            # Totals row
            total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            total_font = Font(bold=True, size=10)
            ws.cell(row=row, column=1, value="TOTAL").border = border
            ws.cell(row=row, column=1).font = total_font
            ws.cell(row=row, column=1).fill = total_fill
            ws.cell(row=row, column=2, value="").border = border
            ws.cell(row=row, column=2).fill = total_fill
            bat_sum = sum(pd["batting"] for _, pd in played)
            bowl_sum = sum(pd["bowling"] for _, pd in played)
            field_sum = sum(pd["fielding"] for _, pd in played)
            gen_sum = sum(pd["generic"] for _, pd in played)
            base_sum = sum(pd["base_total"] for _, pd in played)
            total_sum = data["total"]
            for ci, val in enumerate([bat_sum, bowl_sum, field_sum, gen_sum, base_sum, total_sum], 3):
                cell = ws.cell(row=row, column=ci, value=round(val, 1))
                cell.number_format = '0.0'
                cell.alignment = right_align
                cell.border = border
                cell.font = total_font
                cell.fill = total_fill
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        for col in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 12
    
    wb.save(filename)
    print(f"✓ Detailed view exported to {filename}")


def export_master_to_excel(filename: str = "master_scoresheet.xlsx"):
    """Export the cumulative master scoresheet to Excel with two sheets:
       1. Leaderboard – team standings with per-match totals
       2. One sheet per team – player-by-player breakdown across all matches
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    master = load_master_scoresheet()
    match_list = master.get("match_list", [])
    teams = master.get("teams", {})

    if not match_list or not teams:
        print("\nNo data in master scoresheet. Process matches first.\n")
        return

    wb = Workbook()

    # ── Styles ──
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    team_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    team_font = Font(bold=True, color="FFFFFF", size=11)
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    num_fmt = '0.0'

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 1: Leaderboard
    # ═══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Leaderboard"

    ws['A1'] = "IPL Fantasy Premier League 2026 – Master Scoresheet"
    ws['A1'].font = title_font
    num_matches = len(match_list)
    last_col = 3 + num_matches  # Rank, Team, match labels, Total
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    ws['A2'] = f"Matches processed: {num_matches}"
    ws['A2'].font = Font(italic=True, size=10)

    match_labels = []
    for m in match_list:
        match_labels.append(_match_label_for_master_entry(m))

    row = 4
    # Headers: Rank | Team | <match names> | Total
    headers = ["Rank", "Team"] + match_labels + ["Total"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    row += 1

    sorted_teams = sorted(teams.items(),
                          key=lambda x: x[1]["cumulative_total"], reverse=True)

    for rank, (owner, tdata) in enumerate(sorted_teams, 1):
        ws.cell(row=row, column=1, value=rank).border = border
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2, value=owner).border = border
        ws.cell(row=row, column=2).alignment = left

        # Podium highlight
        if rank == 1:
            ws.cell(row=row, column=1).fill = gold_fill
            ws.cell(row=row, column=2).fill = gold_fill
        elif rank == 2:
            ws.cell(row=row, column=1).fill = silver_fill
            ws.cell(row=row, column=2).fill = silver_fill
        elif rank == 3:
            ws.cell(row=row, column=1).fill = bronze_fill
            ws.cell(row=row, column=2).fill = bronze_fill

        for mi, m in enumerate(match_list):
            mid = str(m["match_id"])
            match_total = sum(
                ps["match_scores"].get(mid, {}).get("total", 0)
                for ps in tdata["players"].values()
            )
            col_idx = 3 + mi
            cell = ws.cell(row=row, column=col_idx, value=round(match_total, 1))
            cell.number_format = num_fmt
            cell.alignment = right
            cell.border = border

        total_col = 3 + num_matches
        cell = ws.cell(row=row, column=total_col, value=round(tdata["cumulative_total"], 1))
        cell.number_format = num_fmt
        cell.alignment = right
        cell.border = border
        cell.font = Font(bold=True)
        row += 1

    # Column totals row (sum of all teams per match + grand total)
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    total_font = Font(bold=True, size=10)
    ws.cell(row=row, column=1, value="").border = border
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=2, value="TOTAL").border = border
    ws.cell(row=row, column=2).font = total_font
    ws.cell(row=row, column=2).fill = total_fill
    ws.cell(row=row, column=2).alignment = left
    for mi, m in enumerate(match_list):
        mid = str(m["match_id"])
        col_total = round(sum(
            sum(ps["match_scores"].get(mid, {}).get("total", 0)
                for ps in tdata["players"].values())
            for tdata in teams.values()
        ), 1)
        col_idx = 3 + mi
        cell = ws.cell(row=row, column=col_idx, value=col_total)
        cell.number_format = num_fmt
        cell.alignment = right
        cell.border = border
        cell.font = total_font
        cell.fill = total_fill
    grand_total = round(sum(td["cumulative_total"] for td in teams.values()), 1)
    cell = ws.cell(row=row, column=total_col, value=grand_total)
    cell.number_format = num_fmt
    cell.alignment = right
    cell.border = border
    cell.font = total_font
    cell.fill = total_fill
    row += 1

    # Match legend
    row += 1
    ws.cell(row=row, column=1, value="Match Key:").font = Font(bold=True, size=10)
    row += 1
    for i, m in enumerate(match_list):
        match_label = _match_label_for_master_entry(m)
        if match_label and match_label != m.get("description", ""):
            ws.cell(row=row, column=1, value=f"M{i+1}: {match_label} — {m['description']} (ID: {m['match_id']})")
        else:
            ws.cell(row=row, column=1, value=f"M{i+1}: {m['description']} (ID: {m['match_id']})")
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 16
    for ci in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    # ═══════════════════════════════════════════════════════════════════════
    # One sheet per team
    # ═══════════════════════════════════════════════════════════════════════
    for rank, (owner, tdata) in enumerate(sorted_teams, 1):
        ws = wb.create_sheet(title=f"{rank}. {owner}"[:31])

        ws['A1'] = f"{owner} — Cumulative: {tdata['cumulative_total']} pts"
        ws['A1'].font = Font(bold=True, size=12)
        # Cols: Player | Role | M1 | M2 | ... | Batting | Bowling | Fielding | Generic | Total | MP
        num_extra = 6  # Bat, Bowl, Fld, Gen, Total, MP
        total_cols = 2 + num_matches + num_extra
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

        row = 3
        headers = ["Player", "Role"] + match_labels
        headers.extend(["Batting", "Bowling", "Fielding", "Generic", "Total", "MP"])

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        row += 1

        # Sort players: played first by total desc, then unplayed alpha
        players = list(tdata["players"].items())
        played = [(n, p) for n, p in players if p["cumulative"]["matches_played"] > 0]
        unplayed = [(n, p) for n, p in players if p["cumulative"]["matches_played"] == 0]
        played.sort(key=lambda x: x[1]["cumulative"]["total"], reverse=True)
        unplayed.sort(key=lambda x: x[0])

        for pname, pdata in played + unplayed:
            tag = ""
            if pdata.get("is_captain"):
                tag = " (C)"
            elif pdata.get("is_vice_captain"):
                tag = " (VC)"

            ws.cell(row=row, column=1, value=f"{pname}{tag}").border = border
            ws.cell(row=row, column=1).alignment = left
            ws.cell(row=row, column=2, value=pdata["role"]).border = border
            ws.cell(row=row, column=2).alignment = left

            # Per-match scores
            for mi, m in enumerate(match_list):
                mid = str(m["match_id"])
                ms = pdata["match_scores"].get(mid)
                col_idx = 3 + mi
                if ms and (ms.get("played") or ms.get("total", 0) != 0):
                    cell = ws.cell(row=row, column=col_idx, value=round(ms["total"], 1))
                    cell.number_format = num_fmt
                else:
                    cell = ws.cell(row=row, column=col_idx, value="—")
                cell.alignment = right
                cell.border = border

            # Cumulative columns
            cum = pdata["cumulative"]
            base_col = 3 + num_matches
            if cum["matches_played"] > 0 or cum["total"] != 0:
                for ci, key in enumerate(["batting", "bowling", "fielding", "generic", "total"]):
                    cell = ws.cell(row=row, column=base_col + ci, value=round(cum[key], 1))
                    cell.number_format = num_fmt
                    cell.alignment = right
                    cell.border = border
                ws.cell(row=row, column=base_col + 4).font = Font(bold=True)
                ws.cell(row=row, column=base_col + 5, value=cum["matches_played"]).border = border
                ws.cell(row=row, column=base_col + 5).alignment = center
            else:
                for ci in range(5):
                    cell = ws.cell(row=row, column=base_col + ci, value="—")
                    cell.alignment = center
                    cell.border = border
                ws.cell(row=row, column=base_col + 5, value=0).border = border
                ws.cell(row=row, column=base_col + 5).alignment = center

            row += 1

        # Totals row
        total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        total_font_s = Font(bold=True, size=10)
        ws.cell(row=row, column=1, value="TOTAL").border = border
        ws.cell(row=row, column=1).font = total_font_s
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=2, value="").border = border
        ws.cell(row=row, column=2).fill = total_fill
        for mi, m in enumerate(match_list):
            mid = str(m["match_id"])
            col_sum = round(sum(
                pdata["match_scores"].get(mid, {}).get("total", 0)
                for pdata in tdata["players"].values()
            ), 1)
            col_idx = 3 + mi
            cell = ws.cell(row=row, column=col_idx, value=col_sum)
            cell.number_format = num_fmt
            cell.alignment = right
            cell.border = border
            cell.font = total_font_s
            cell.fill = total_fill
        for ci_off, key in enumerate(["batting", "bowling", "fielding", "generic", "total"]):
            col_idx = base_col + ci_off
            cum_sum = round(sum(
                pdata["cumulative"].get(key, 0)
                for pdata in tdata["players"].values()
            ), 1)
            cell = ws.cell(row=row, column=col_idx, value=cum_sum)
            cell.number_format = num_fmt
            cell.alignment = right
            cell.border = border
            cell.font = total_font_s
            cell.fill = total_fill
        # MP column
        mp_sum = sum(pdata["cumulative"].get("matches_played", 0)
                     for pdata in tdata["players"].values())
        ws.cell(row=row, column=base_col + 5, value=mp_sum).border = border
        ws.cell(row=row, column=base_col + 5).font = total_font_s
        ws.cell(row=row, column=base_col + 5).fill = total_fill
        ws.cell(row=row, column=base_col + 5).alignment = center
        row += 1

        # Column widths
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 14
        for ci in range(3, total_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 12

    wb.save(filename)
    print(f"✓ Master scoresheet exported to {filename}")


# ══════════════════════════════════════════════════════════════════════════════
# MATCH SCORECARD – raw cricket scorecard display
# ══════════════════════════════════════════════════════════════════════════════

def show_scorecard(scorecard_data: dict, match_id: int = 0):
    """Print the cricket scorecard (batting & bowling) for each innings."""
    innings_list = scorecard_data.get("scorecard", [])
    status = scorecard_data.get("status", "")
    seo = scorecard_data.get("appindex", {}).get("seotitle", "")
    title = seo.split("|")[0].strip() if seo else f"Match {match_id}"

    print(f"\n{'='*75}")
    print(f"  SCORECARD – {title}")
    print(f"{'='*75}")
    if status:
        print(f"  Result: {status}")

    for inn in innings_list:
        team = inn.get("batteamname", "Unknown")
        short = inn.get("batteamsname", "")
        score = inn.get("score", 0)
        wickets = inn.get("wickets", 0)
        overs = inn.get("overs", 0)
        rr = inn.get("runrate", "0")
        extras = inn.get("extras", 0)
        if isinstance(extras, dict):
            extras_total = extras.get("total", 0)
            extras_detail = ", ".join(
                f"{k} {v}" for k, v in extras.items()
                if k != "total" and v > 0
            )
        else:
            extras_total = extras
            extras_detail = ""

        inn_id = inn.get("inningsid", "")
        print(f"\n{'━'*75}")
        print(f"  Innings {inn_id}: {team} ({short}) — {score}/{wickets} in {overs} overs (RR: {rr})")
        print(f"{'━'*75}")

        # Batting
        bat_rows = []
        for b in inn.get("batsman", []):
            not_out = "not out" in b.get("outdec", "").lower()
            dismissal = b.get("outdec", "")
            if dismissal.lower() == "not out":
                dismissal = "not out"
            name = b.get("name", "")
            tags = []
            if b.get("iscaptain"):
                tags.append("c")
            if b.get("iskeeper"):
                tags.append("wk")
            if tags:
                name += f" ({'/'.join(tags)})"
            bat_rows.append([
                name, dismissal,
                b.get("runs", 0), b.get("balls", 0),
                b.get("fours", 0), b.get("sixes", 0),
                b.get("strkrate", "0"),
            ])

        print("\n  BATTING:")
        print(tabulate(
            bat_rows,
            headers=["Batter", "Dismissal", "R", "B", "4s", "6s", "SR"],
            tablefmt="simple", numalign="right",
            colalign=("left", "left", "right", "right", "right", "right", "right"),
        ))

        extras_str = f"  Extras: {extras_total}"
        if extras_detail:
            extras_str += f" ({extras_detail})"
        print(f"\n{extras_str}")
        print(f"  TOTAL: {score}/{wickets} ({overs} Ov)")

        # Bowling
        bowl_rows = []
        for bw in inn.get("bowler", []):
            name = bw.get("name", "")
            bowl_rows.append([
                name,
                bw.get("overs", 0), bw.get("maidens", 0),
                bw.get("runs", 0), bw.get("wickets", 0),
                bw.get("economy", "0"),
            ])

        print("\n  BOWLING:")
        print(tabulate(
            bowl_rows,
            headers=["Bowler", "O", "M", "R", "W", "Econ"],
            tablefmt="simple", numalign="right",
            colalign=("left", "right", "right", "right", "right", "right"),
        ))

    print()


def export_scorecard_to_excel(scorecard_data: dict, filename: str,
                               match_id: int = 0):
    """Export the cricket scorecard to Excel with one sheet per innings."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    innings_list = scorecard_data.get("scorecard", [])
    status = scorecard_data.get("status", "")
    seo = scorecard_data.get("appindex", {}).get("seotitle", "")
    title = seo.split("|")[0].strip() if seo else f"Match {match_id}"

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    team_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    team_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    section_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    num_fmt = '0.00'

    for inn in innings_list:
        team = inn.get("batteamname", "Unknown")
        short = inn.get("batteamsname", "")
        score = inn.get("score", 0)
        wickets = inn.get("wickets", 0)
        overs = inn.get("overs", 0)
        rr = inn.get("runrate", "0")
        extras = inn.get("extras", 0)
        if isinstance(extras, dict):
            extras_total = extras.get("total", 0)
        else:
            extras_total = extras
        inn_id = inn.get("inningsid", "")

        ws = wb.create_sheet(title=f"{short} Innings"[:31])

        short_match = _short_match_name_from_scorecard(scorecard_data, match_id)
        if short_match and short_match != title:
            ws['A1'] = short_match
            ws['A1'].font = Font(bold=True, size=13)
            ws.merge_cells('A1:G1')
            ws['A2'] = title
            ws['A2'].font = Font(italic=True, size=10, color="555555")
            ws.merge_cells('A2:G2')
            row = 4
            if status and inn_id == 1:
                ws['A3'] = status
                ws['A3'].font = Font(italic=True, size=10, color="555555")
                ws.merge_cells('A3:G3')
                row = 5
        else:
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=13)
            ws.merge_cells('A1:G1')
            row = 4
            if status and inn_id == 1:
                ws['A2'] = status
                ws['A2'].font = Font(italic=True, size=10, color="555555")
                ws.merge_cells('A2:G2')
                row = 5

        # ── Batting ──
        ws.cell(row=row, column=1, value="BATTING").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = section_fill
        row += 1

        bat_headers = ["Batter", "Dismissal", "R", "B", "4s", "6s", "SR"]
        for ci, h in enumerate(bat_headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        row += 1

        batsmen = inn.get("batsman", [])
        for b in batsmen:
            name = b.get("name", "")
            tags = []
            if b.get("iscaptain"):
                tags.append("c")
            if b.get("iskeeper"):
                tags.append("wk")
            if tags:
                name += f" ({'/'.join(tags)})"

            dismissal = b.get("outdec", "")
            ws.cell(row=row, column=1, value=name).border = border
            ws.cell(row=row, column=1).alignment = left_align
            ws.cell(row=row, column=2, value=dismissal).border = border
            ws.cell(row=row, column=2).alignment = left_align
            ws.cell(row=row, column=3, value=b.get("runs", 0)).border = border
            ws.cell(row=row, column=3).alignment = right_align
            ws.cell(row=row, column=4, value=b.get("balls", 0)).border = border
            ws.cell(row=row, column=4).alignment = right_align
            ws.cell(row=row, column=5, value=b.get("fours", 0)).border = border
            ws.cell(row=row, column=5).alignment = right_align
            ws.cell(row=row, column=6, value=b.get("sixes", 0)).border = border
            ws.cell(row=row, column=6).alignment = right_align

            sr = b.get("strkrate", "0")
            try:
                sr_val = float(sr)
            except (ValueError, TypeError):
                sr_val = 0
            ws.cell(row=row, column=7, value=sr_val).border = border
            ws.cell(row=row, column=7).alignment = right_align
            ws.cell(row=row, column=7).number_format = num_fmt
            row += 1

        # Batting column totals
        total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        total_font = Font(bold=True, size=10)
        ws.cell(row=row, column=1, value="TOTAL").border = border
        ws.cell(row=row, column=1).font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=2, value="").border = border
        ws.cell(row=row, column=2).fill = total_fill
        for ci, key in enumerate(["runs", "balls", "fours", "sixes"], 3):
            col_sum = sum(b.get(key, 0) for b in batsmen)
            cell = ws.cell(row=row, column=ci, value=col_sum)
            cell.alignment = right_align
            cell.border = border
            cell.font = total_font
            cell.fill = total_fill
        ws.cell(row=row, column=7, value="").border = border
        ws.cell(row=row, column=7).fill = total_fill
        row += 1

        # Extras & Total
        ws.cell(row=row, column=1, value="Extras").font = Font(italic=True)
        ws.cell(row=row, column=3, value=extras_total).font = Font(italic=True)
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, size=11)
        ws.cell(row=row, column=3, value=f"{score}/{wickets} ({overs} Ov)").font = Font(bold=True)
        row += 2

        # ── Bowling ──
        ws.cell(row=row, column=1, value="BOWLING").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = section_fill
        row += 1

        bowl_headers = ["Bowler", "O", "M", "R", "W", "Econ"]
        for ci, h in enumerate(bowl_headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        row += 1

        bowlers = inn.get("bowler", [])
        for bw in bowlers:
            ws.cell(row=row, column=1, value=bw.get("name", "")).border = border
            ws.cell(row=row, column=1).alignment = left_align
            ws.cell(row=row, column=2, value=bw.get("overs", 0)).border = border
            ws.cell(row=row, column=2).alignment = right_align
            ws.cell(row=row, column=3, value=bw.get("maidens", 0)).border = border
            ws.cell(row=row, column=3).alignment = right_align
            ws.cell(row=row, column=4, value=bw.get("runs", 0)).border = border
            ws.cell(row=row, column=4).alignment = right_align
            ws.cell(row=row, column=5, value=bw.get("wickets", 0)).border = border
            ws.cell(row=row, column=5).alignment = right_align

            econ = bw.get("economy", "0")
            try:
                econ_val = float(econ)
            except (ValueError, TypeError):
                econ_val = 0
            ws.cell(row=row, column=6, value=econ_val).border = border
            ws.cell(row=row, column=6).alignment = right_align
            ws.cell(row=row, column=6).number_format = num_fmt
            row += 1

        # Bowling column totals
        ws.cell(row=row, column=1, value="TOTAL").border = border
        ws.cell(row=row, column=1).font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        overs_sum = round(sum(float(bw.get("overs", 0)) for bw in bowlers), 1)
        for ci, key in enumerate(["maidens", "runs", "wickets"], 3):
            col_sum = sum(bw.get(key, 0) for bw in bowlers)
            cell = ws.cell(row=row, column=ci, value=col_sum)
            cell.alignment = right_align
            cell.border = border
            cell.font = total_font
            cell.fill = total_fill
        cell = ws.cell(row=row, column=2, value=overs_sum)
        cell.alignment = right_align
        cell.border = border
        cell.font = total_font
        cell.fill = total_fill
        ws.cell(row=row, column=6, value="").border = border
        ws.cell(row=row, column=6).fill = total_fill
        row += 1

        # Column widths
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 10

    wb.save(filename)
    print(f"✓ Scorecard exported to {filename}")


# ══════════════════════════════════════════════════════════════════════════════
# TEAM POINTS PER MATCH – fantasy team points breakdown for a single match
# ══════════════════════════════════════════════════════════════════════════════

def export_team_points_to_excel(match_results: dict, description: str = "",
                                 filename: str = "team_points.xlsx"):
    """Export matchwise team points to Excel.

    Sheet 1: Team Rankings – all teams ranked by match points.
    One sheet per team: player-by-player breakdown for that match.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── Styles ──
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    team_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    team_font = Font(bold=True, color="FFFFFF", size=11)
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
    played_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    num_fmt = '0.0'

    sorted_teams = sorted(match_results.items(),
                          key=lambda x: x[1]["total"], reverse=True)

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 1: Team Rankings
    # ═══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Team Rankings"

    ws['A1'] = f"IPL Fantasy Premier League 2026 – Team Points"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')

    short_description = _short_match_name_from_description(description) if description else ""
    if short_description and short_description != description:
        ws['A2'] = short_description
        ws['A2'].font = Font(italic=True, size=10, color="555555")
        ws.merge_cells('A2:G2')
        ws['A3'] = description
        ws['A3'].font = Font(italic=True, size=10, color="555555")
        ws.merge_cells('A3:G3')
        row = 5
    elif description:
        ws['A2'] = description
        ws['A2'].font = Font(italic=True, size=10, color="555555")
        ws.merge_cells('A2:G2')
        row = 4
    else:
        row = 4
    headers = ["Rank", "Team", "Batting", "Bowling", "Fielding", "Generic", "Total"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    row += 1

    for rank, (owner, data) in enumerate(sorted_teams, 1):
        # Aggregate category totals across all players
        bat_total = sum(p["batting"] for p in data["players"].values())
        bowl_total = sum(p["bowling"] for p in data["players"].values())
        field_total = sum(p["fielding"] for p in data["players"].values())
        gen_total = sum(p["generic"] for p in data["players"].values())

        ws.cell(row=row, column=1, value=rank).border = border
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2, value=owner).border = border
        ws.cell(row=row, column=2).alignment = left_align

        # Podium highlights
        if rank == 1:
            ws.cell(row=row, column=1).fill = gold_fill
            ws.cell(row=row, column=2).fill = gold_fill
        elif rank == 2:
            ws.cell(row=row, column=1).fill = silver_fill
            ws.cell(row=row, column=2).fill = silver_fill
        elif rank == 3:
            ws.cell(row=row, column=1).fill = bronze_fill
            ws.cell(row=row, column=2).fill = bronze_fill

        for ci, val in enumerate([bat_total, bowl_total, field_total, gen_total, data["total"]], 3):
            cell = ws.cell(row=row, column=ci, value=round(val, 1))
            cell.number_format = num_fmt
            cell.alignment = right_align
            cell.border = border
        ws.cell(row=row, column=7).font = Font(bold=True)
        row += 1

    # Column totals row for the Rankings sheet
    total_fill_r = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    total_font_r = Font(bold=True, size=10)
    ws.cell(row=row, column=1, value="").border = border
    ws.cell(row=row, column=1).fill = total_fill_r
    ws.cell(row=row, column=2, value="TOTAL").border = border
    ws.cell(row=row, column=2).font = total_font_r
    ws.cell(row=row, column=2).fill = total_fill_r
    ws.cell(row=row, column=2).alignment = left_align
    all_players_flat = [p for _, d in match_results.items() for p in d["players"].values()]
    for ci, key in enumerate(["batting", "bowling", "fielding", "generic"], 3):
        col_sum = round(sum(p[key] for p in all_players_flat), 1)
        cell = ws.cell(row=row, column=ci, value=col_sum)
        cell.number_format = num_fmt
        cell.alignment = right_align
        cell.border = border
        cell.font = total_font_r
        cell.fill = total_fill_r
    grand_total = round(sum(d["total"] for _, d in match_results.items()), 1)
    cell = ws.cell(row=row, column=7, value=grand_total)
    cell.number_format = num_fmt
    cell.alignment = right_align
    cell.border = border
    cell.font = total_font_r
    cell.fill = total_fill_r
    row += 1

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 16
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 12

    # ═══════════════════════════════════════════════════════════════════════
    # One sheet per team – player breakdown
    # ═══════════════════════════════════════════════════════════════════════
    for rank, (owner, data) in enumerate(sorted_teams, 1):
        ws = wb.create_sheet(title=f"{rank}. {owner}"[:31])

        ws['A1'] = f"{owner} — Total: {data['total']} pts"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:H1')

        if description:
            ws['A2'] = description
            ws['A2'].font = Font(italic=True, size=10, color="555555")
            ws.merge_cells('A2:H2')

        row = 4
        headers = ["Player", "Role", "Batting", "Bowling", "Fielding", "Generic", "Base", "Total"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        row += 1

        players = data.get("players", {})
        played = [(n, p) for n, p in players.items() if p.get("played")]
        not_played = [(n, p) for n, p in players.items() if not p.get("played")]
        played.sort(key=lambda x: x[1]["total"], reverse=True)
        not_played.sort(key=lambda x: x[0])

        for pname, pd in played:
            tag = _player_tag(pd)
            ws.cell(row=row, column=1, value=f"{pname}{tag}").border = border
            ws.cell(row=row, column=1).alignment = left_align
            ws.cell(row=row, column=2, value=pd.get("role", "")).border = border
            ws.cell(row=row, column=2).alignment = left_align

            for ci, key in enumerate(["batting", "bowling", "fielding", "generic", "base_total", "total"], 3):
                cell = ws.cell(row=row, column=ci, value=round(pd[key], 1))
                cell.number_format = num_fmt
                cell.alignment = right_align
                cell.border = border

            # Highlight played rows
            for ci in range(1, 9):
                ws.cell(row=row, column=ci).fill = played_fill

            ws.cell(row=row, column=8).font = Font(bold=True)

            # Captain/VC marker highlight
            if pd.get("multiplier", 1.0) == 1.5:
                ws.cell(row=row, column=1).font = Font(bold=True, color="FF0000")
            elif pd.get("multiplier", 1.0) == 1.25:
                ws.cell(row=row, column=1).font = Font(bold=True, color="0070C0")

            row += 1

        for pname, pd in not_played:
            ws.cell(row=row, column=1, value=pname).border = border
            ws.cell(row=row, column=1).alignment = left_align
            ws.cell(row=row, column=2, value=pd.get("role", "")).border = border
            ws.cell(row=row, column=2).alignment = left_align
            for ci in range(3, 9):
                cell = ws.cell(row=row, column=ci, value="—")
                cell.alignment = center
                cell.border = border
            row += 1

        # Team total row
        row += 1
        ws.cell(row=row, column=1, value="TEAM TOTAL").font = Font(bold=True, size=11)
        ws.cell(row=row, column=1).border = border
        bat_total = sum(p["batting"] for p in players.values())
        bowl_total = sum(p["bowling"] for p in players.values())
        field_total = sum(p["fielding"] for p in players.values())
        gen_total = sum(p["generic"] for p in players.values())
        base_total = sum(p["base_total"] for p in players.values())
        for ci, val in enumerate([bat_total, bowl_total, field_total, gen_total, base_total, data["total"]], 3):
            cell = ws.cell(row=row, column=ci, value=round(val, 1))
            cell.number_format = num_fmt
            cell.alignment = right_align
            cell.border = border
            cell.font = Font(bold=True)

        # Column widths
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 14
        for col in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 12

    wb.save(filename)
    print(f"✓ Team points exported to {filename}")


# ══════════════════════════════════════════════════════════════════════════════
# TEAM DATA EXPORT – all fantasy teams with players, roles, prices
# ══════════════════════════════════════════════════════════════════════════════

def export_teams_to_excel(filename: str = "teams.xlsx"):
    """Export all fantasy team rosters to Excel.

    Sheet 1: Overview – all teams with budget, captain, VC, player count, total spent.
    One sheet per team: full player roster with role, country, price.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from storage import load_teams

    teams_data = load_teams()
    teams = teams_data.get("teams", {})

    if not teams:
        print("\nNo teams found.\n")
        return

    wb = Workbook()

    # ── Styles ──
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    captain_font = Font(bold=True, color="FF0000")
    vc_font = Font(bold=True, color="0070C0")
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 1: Overview
    # ═══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Overview"

    ws['A1'] = "IPL Fantasy Premier League 2026 – Team Rosters"
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')

    row = 3
    headers = ["#", "Owner", "Captain", "Vice Captain", "Players",
               "Indian", "Overseas", "Budget", "Spent", "Remaining"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    row += 1

    # Sort teams by total spent descending
    team_list = []
    for owner, team in teams.items():
        players = team.get("players", [])
        spent = sum(p.get("price", 0) for p in players)
        indian = sum(1 for p in players if p.get("country") == "Indian")
        overseas = sum(1 for p in players if p.get("country") == "Overseas")
        team_list.append((owner, team, len(players), indian, overseas, spent))
    team_list.sort(key=lambda x: x[5], reverse=True)

    for rank, (owner, team, count, indian, overseas, spent) in enumerate(team_list, 1):
        budget = team.get("budget", 9000)
        remaining = budget - spent

        ws.cell(row=row, column=1, value=rank).border = border
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2, value=owner).border = border
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=3, value=team.get("captain", "")).border = border
        ws.cell(row=row, column=3).alignment = left_align
        ws.cell(row=row, column=4, value=team.get("vice_captain", "")).border = border
        ws.cell(row=row, column=4).alignment = left_align
        ws.cell(row=row, column=5, value=count).border = border
        ws.cell(row=row, column=5).alignment = center
        ws.cell(row=row, column=6, value=indian).border = border
        ws.cell(row=row, column=6).alignment = center
        ws.cell(row=row, column=7, value=overseas).border = border
        ws.cell(row=row, column=7).alignment = center
        ws.cell(row=row, column=8, value=budget).border = border
        ws.cell(row=row, column=8).alignment = right_align
        ws.cell(row=row, column=9, value=spent).border = border
        ws.cell(row=row, column=9).alignment = right_align
        ws.cell(row=row, column=10, value=remaining).border = border
        ws.cell(row=row, column=10).alignment = right_align

        if rank <= 3:
            fill = [gold_fill, silver_fill, bronze_fill][rank - 1]
            ws.cell(row=row, column=1).fill = fill
            ws.cell(row=row, column=2).fill = fill

        row += 1

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12

    # ═══════════════════════════════════════════════════════════════════════
    # One sheet per team
    # ═══════════════════════════════════════════════════════════════════════
    for rank, (owner, team, count, indian, overseas, spent) in enumerate(team_list, 1):
        ws = wb.create_sheet(title=f"{rank}. {owner}"[:31])
        budget = team.get("budget", 9000)
        captain = team.get("captain", "")
        vice_captain = team.get("vice_captain", "")

        ws['A1'] = f"{owner} — {count} players, Spent: {spent}, Remaining: {budget - spent}"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')

        ws['A2'] = f"Captain: {captain}  |  Vice Captain: {vice_captain}"
        ws['A2'].font = Font(italic=True, size=10, color="555555")
        ws.merge_cells('A2:F2')

        row = 4
        headers = ["#", "Player", "Role", "Country", "Price", "Tag"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        row += 1

        players = team.get("players", [])
        players_sorted = sorted(players, key=lambda p: p.get("price", 0), reverse=True)

        for i, p in enumerate(players_sorted, 1):
            name = p.get("name", "")
            tag = ""
            if name == captain:
                tag = "C"
            elif name == vice_captain:
                tag = "VC"

            ws.cell(row=row, column=1, value=i).border = border
            ws.cell(row=row, column=1).alignment = center
            ws.cell(row=row, column=2, value=name).border = border
            ws.cell(row=row, column=2).alignment = left_align
            ws.cell(row=row, column=3, value=p.get("role", "")).border = border
            ws.cell(row=row, column=3).alignment = left_align
            ws.cell(row=row, column=4, value=p.get("country", "")).border = border
            ws.cell(row=row, column=4).alignment = center
            ws.cell(row=row, column=5, value=p.get("price", 0)).border = border
            ws.cell(row=row, column=5).alignment = right_align
            ws.cell(row=row, column=6, value=tag).border = border
            ws.cell(row=row, column=6).alignment = center

            if tag == "C":
                ws.cell(row=row, column=2).font = captain_font
                ws.cell(row=row, column=6).font = captain_font
            elif tag == "VC":
                ws.cell(row=row, column=2).font = vc_font
                ws.cell(row=row, column=6).font = vc_font

            row += 1

        # Total row
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=5, value=spent).font = Font(bold=True)
        ws.cell(row=row, column=5).alignment = right_align

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 8

    wb.save(filename)
    print(f"✓ Teams exported to {filename}")


# ══════════════════════════════════════════════════════════════════════════════
# ANALYTICS EXPORT – key stats & metrics for infographic display
# ══════════════════════════════════════════════════════════════════════════════

def export_analytics_to_excel(filename: str = "analytics.xlsx"):
    """Export analytics dashboard to Excel with multiple sheets:
       1. Top Players by Role – highest scorers per role (Batsman, Bowler, WK, AR)
       2. Top Players Overall – overall highest scorers regardless of role
       3. Points Breakdown – batting vs bowling vs fielding contribution
       4. Best Single Match – best individual match performances
       5. Team Analytics – team-level stats
       6. Category Leaders – top batting, bowling, fielding point earners
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    master = load_master_scoresheet()
    match_list = master.get("match_list", [])
    teams = master.get("teams", {})

    if not match_list or not teams:
        print("\nNo data in master scoresheet. Process matches first.\n")
        return

    wb = Workbook()

    # ── Styles ──
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11, color="366092")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
    podium_fills = [gold_fill, silver_fill, bronze_fill]
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    section_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    right_a = Alignment(horizontal="right", vertical="center")
    left_a = Alignment(horizontal="left", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    num_fmt = '0.0'
    bold = Font(bold=True)

    # ── Gather all player data ──
    all_players = []  # list of dicts with flat player info
    for owner, tdata in teams.items():
        for pname, pdata in tdata["players"].items():
            cum = pdata["cumulative"]
            if cum["matches_played"] == 0:
                continue
            # Best single match
            best_match_pts = 0
            best_match_id = ""
            for mid, ms in pdata["match_scores"].items():
                if ms.get("played") and ms["total"] > best_match_pts:
                    best_match_pts = ms["total"]
                    best_match_id = mid
            all_players.append({
                "name": pname,
                "owner": owner,
                "role": pdata["role"],
                "is_captain": pdata.get("is_captain", False),
                "is_vice_captain": pdata.get("is_vice_captain", False),
                "batting": cum["batting"],
                "bowling": cum["bowling"],
                "fielding": cum["fielding"],
                "generic": cum["generic"],
                "total": cum["total"],
                "matches_played": cum["matches_played"],
                "avg_per_match": round(cum["total"] / cum["matches_played"], 1),
                "best_match_pts": best_match_pts,
                "best_match_id": best_match_id,
            })

    num_matches = len(match_list)

    def _write_headers(ws, row, headers):
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

    def _write_row(ws, row, values, highlight_rank=None):
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = border
            if isinstance(val, (int, float)):
                cell.alignment = right_a
                cell.number_format = num_fmt
            else:
                cell.alignment = left_a
        if highlight_rank is not None and 0 <= highlight_rank < 3:
            ws.cell(row=row, column=1).fill = podium_fills[highlight_rank]

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 1: Top Players by Role
    # ═══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Top by Role"

    ws['A1'] = "IPL Fantasy Premier League 2026 – Top Players by Role"
    ws['A1'].font = title_font
    ws.merge_cells('A1:I1')
    ws['A2'] = f"Across {num_matches} matches"
    ws['A2'].font = Font(italic=True, size=10)

    row = 4
    role_order = ["Batsman", "Wicketkeeper", "Bowler", "Allrounder"]
    headers = ["#", "Player", "Owner", "Total Pts", "Batting", "Bowling", "Fielding", "MP", "Avg/Match"]

    for role in role_order:
        role_players = sorted([p for p in all_players if p["role"] == role],
                              key=lambda x: x["total"], reverse=True)
        if not role_players:
            continue

        # Section header
        ws.cell(row=row, column=1, value=f"TOP {role.upper()}S")
        ws.cell(row=row, column=1).font = subtitle_font
        row += 1
        _write_headers(ws, row, headers)
        row += 1

        for rank, p in enumerate(role_players[:10], 1):
            _write_row(ws, row, [
                rank, p["name"], p["owner"], p["total"],
                p["batting"], p["bowling"], p["fielding"],
                p["matches_played"], p["avg_per_match"],
            ], highlight_rank=rank - 1 if rank <= 3 else None)
            row += 1
        row += 1

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 12
    for col in ['D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 12

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 2: Top Players Overall
    # ═══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="Top Overall")
    ws2['A1'] = "Top Fantasy Players – Overall Rankings"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:J1')

    row = 3
    headers2 = ["#", "Player", "Role", "Owner", "Total Pts", "Batting",
                "Bowling", "Fielding", "MP", "Avg/Match"]
    _write_headers(ws2, row, headers2)
    row += 1

    top_overall = sorted(all_players, key=lambda x: x["total"], reverse=True)
    for rank, p in enumerate(top_overall[:30], 1):
        _write_row(ws2, row, [
            rank, p["name"], p["role"], p["owner"], p["total"],
            p["batting"], p["bowling"], p["fielding"],
            p["matches_played"], p["avg_per_match"],
        ], highlight_rank=rank - 1 if rank <= 3 else None)
        row += 1

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 26
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 12
    for col in ['E', 'F', 'G', 'H', 'I', 'J']:
        ws2.column_dimensions[col].width = 12

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 3: Category Leaders (Best Batting / Bowling / Fielding earners)
    # ═══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet(title="Category Leaders")
    ws3['A1'] = "Category Leaders – Who Earns Most Points in Each Category"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:G1')

    row = 3
    categories = [
        ("TOP BATTING POINT EARNERS", "batting"),
        ("TOP BOWLING POINT EARNERS", "bowling"),
        ("TOP FIELDING POINT EARNERS", "fielding"),
        ("HIGHEST AVERAGE PER MATCH", "avg_per_match"),
    ]
    cat_headers = ["#", "Player", "Role", "Owner", "Category Pts", "Total Pts", "MP"]

    for cat_title, cat_key in categories:
        ws3.cell(row=row, column=1, value=cat_title).font = subtitle_font
        row += 1
        _write_headers(ws3, row, cat_headers)
        row += 1

        sorted_cat = sorted(all_players, key=lambda x: x[cat_key], reverse=True)
        for rank, p in enumerate(sorted_cat[:10], 1):
            _write_row(ws3, row, [
                rank, p["name"], p["role"], p["owner"],
                p[cat_key], p["total"], p["matches_played"],
            ], highlight_rank=rank - 1 if rank <= 3 else None)
            row += 1
        row += 1

    ws3.column_dimensions['A'].width = 6
    ws3.column_dimensions['B'].width = 26
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 12
    for col in ['E', 'F', 'G']:
        ws3.column_dimensions[col].width = 14

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 4: Best Single Match Performances
    # ═══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet(title="Best Match Perf")
    ws4['A1'] = "Best Single-Match Performances"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:H1')

    # Flatten all individual match performances
    match_perfs = []
    match_id_to_desc = {str(m["match_id"]): m["description"] for m in match_list}
    for owner, tdata in teams.items():
        for pname, pdata in tdata["players"].items():
            for mid, ms in pdata["match_scores"].items():
                if ms.get("played") and ms["total"] != 0:
                    match_perfs.append({
                        "name": pname,
                        "role": pdata["role"],
                        "owner": owner,
                        "match": match_id_to_desc.get(mid, f"Match {mid}"),
                        "match_id": mid,
                        "batting": ms["batting"],
                        "bowling": ms["bowling"],
                        "fielding": ms["fielding"],
                        "total": ms["total"],
                    })

    row = 3
    perf_headers = ["#", "Player", "Role", "Owner", "Match", "Batting", "Bowling", "Total"]
    _write_headers(ws4, row, perf_headers)
    row += 1

    top_perfs = sorted(match_perfs, key=lambda x: x["total"], reverse=True)
    for rank, p in enumerate(top_perfs[:30], 1):
        _write_row(ws4, row, [
            rank, p["name"], p["role"], p["owner"], p["match"],
            p["batting"], p["bowling"], p["total"],
        ], highlight_rank=rank - 1 if rank <= 3 else None)
        row += 1

    ws4.column_dimensions['A'].width = 6
    ws4.column_dimensions['B'].width = 26
    ws4.column_dimensions['C'].width = 14
    ws4.column_dimensions['D'].width = 12
    ws4.column_dimensions['E'].width = 30
    for col in ['F', 'G', 'H']:
        ws4.column_dimensions[col].width = 12

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 5: Team Analytics
    # ═══════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet(title="Team Analytics")
    ws5['A1'] = "Team Analytics – Performance Breakdown"
    ws5['A1'].font = title_font
    ws5.merge_cells('A1:L1')

    row = 3
    team_headers = ["#", "Team", "Total Pts", "Batting Pts", "Bowling Pts",
                    "Fielding Pts", "Generic Pts", "Players Used",
                    "Best Player", "Best Pts", "Avg/Match", "Bat%"]
    _write_headers(ws5, row, team_headers)
    row += 1

    sorted_teams = sorted(teams.items(),
                          key=lambda x: x[1]["cumulative_total"], reverse=True)

    for rank, (owner, tdata) in enumerate(sorted_teams, 1):
        # Aggregate team stats
        team_bat = team_bowl = team_fld = team_gen = 0
        players_used = 0
        best_player = ""
        best_pts = 0
        for pname, pdata in tdata["players"].items():
            cum = pdata["cumulative"]
            if cum["matches_played"] > 0:
                players_used += 1
                team_bat += cum["batting"]
                team_bowl += cum["bowling"]
                team_fld += cum["fielding"]
                team_gen += cum["generic"]
                if cum["total"] > best_pts:
                    best_pts = cum["total"]
                    best_player = pname

        total = tdata["cumulative_total"]
        avg_per_match = round(total / num_matches, 1) if num_matches else 0
        bat_pct = round((team_bat / total * 100), 1) if total else 0

        _write_row(ws5, row, [
            rank, owner, round(total, 1), round(team_bat, 1),
            round(team_bowl, 1), round(team_fld, 1), round(team_gen, 1),
            players_used, best_player, round(best_pts, 1),
            avg_per_match, bat_pct,
        ], highlight_rank=rank - 1 if rank <= 3 else None)
        row += 1

    ws5.column_dimensions['A'].width = 6
    ws5.column_dimensions['B'].width = 14
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws5.column_dimensions[col].width = 12
    ws5.column_dimensions['I'].width = 22
    ws5.column_dimensions['J'].width = 10
    ws5.column_dimensions['K'].width = 12
    ws5.column_dimensions['L'].width = 8

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 6: Points Source Distribution
    # ═══════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet(title="Points Distribution")
    ws6['A1'] = "Points Source Distribution – Where Do Points Come From?"
    ws6['A1'].font = title_font
    ws6.merge_cells('A1:H1')

    # Overall distribution
    row = 3
    ws6.cell(row=row, column=1, value="OVERALL POINTS BREAKDOWN").font = subtitle_font
    row += 1
    total_bat = sum(p["batting"] for p in all_players)
    total_bowl = sum(p["bowling"] for p in all_players)
    total_fld = sum(p["fielding"] for p in all_players)
    total_gen = sum(p["generic"] for p in all_players)
    grand_total = total_bat + total_bowl + total_fld + total_gen

    dist_headers = ["Category", "Total Points", "Percentage"]
    _write_headers(ws6, row, dist_headers)
    row += 1

    for cat, val in [("Batting", total_bat), ("Bowling", total_bowl),
                     ("Fielding", total_fld), ("Generic (XI + MoM)", total_gen)]:
        pct = round(val / grand_total * 100, 1) if grand_total else 0
        _write_row(ws6, row, [cat, round(val, 1), f"{pct}%"])
        row += 1
    _write_row(ws6, row, ["TOTAL", round(grand_total, 1), "100%"])
    ws6.cell(row=row, column=1).font = bold
    ws6.cell(row=row, column=2).font = bold
    row += 2

    # Per-role distribution
    ws6.cell(row=row, column=1, value="POINTS BY ROLE").font = subtitle_font
    row += 1
    role_headers = ["Role", "Players", "Total Pts", "Batting", "Bowling",
                    "Fielding", "Avg/Player", "% of All Pts"]
    _write_headers(ws6, row, role_headers)
    row += 1

    for role in role_order:
        rp = [p for p in all_players if p["role"] == role]
        if not rp:
            continue
        r_total = sum(p["total"] for p in rp)
        r_bat = sum(p["batting"] for p in rp)
        r_bowl = sum(p["bowling"] for p in rp)
        r_fld = sum(p["fielding"] for p in rp)
        r_avg = round(r_total / len(rp), 1) if rp else 0
        r_pct = round(r_total / grand_total * 100, 1) if grand_total else 0
        _write_row(ws6, row, [
            role, len(rp), round(r_total, 1), round(r_bat, 1),
            round(r_bowl, 1), round(r_fld, 1), r_avg, f"{r_pct}%",
        ])
        row += 1

    ws6.column_dimensions['A'].width = 20
    ws6.column_dimensions['B'].width = 14
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws6.column_dimensions[col].width = 14

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 7: MVP & Records
    # ═══════════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet(title="MVP & Records")
    ws7['A1'] = "MVP & Key Records"
    ws7['A1'].font = title_font
    ws7.merge_cells('A1:D1')

    row = 3
    record_font = Font(bold=True, size=11, color="366092")
    value_font = Font(bold=True, size=12)

    # Find records
    if all_players:
        mvp = max(all_players, key=lambda x: x["total"])
        best_batter = max(all_players, key=lambda x: x["batting"])
        best_bowler = max(all_players, key=lambda x: x["bowling"])
        best_fielder = max(all_players, key=lambda x: x["fielding"])
        best_avg = max(all_players, key=lambda x: x["avg_per_match"])
        best_single = max(all_players, key=lambda x: x["best_match_pts"])
        most_matches = max(all_players, key=lambda x: x["matches_played"])

    records = [
        ("Overall MVP", mvp["name"], f"{mvp['total']} pts", f"({mvp['role']}, {mvp['owner']})"),
        ("Best Batter (by Fantasy Pts)", best_batter["name"], f"{best_batter['batting']} batting pts", f"({best_batter['role']}, {best_batter['owner']})"),
        ("Best Bowler (by Fantasy Pts)", best_bowler["name"], f"{best_bowler['bowling']} bowling pts", f"({best_bowler['role']}, {best_bowler['owner']})"),
        ("Best Fielder (by Fantasy Pts)", best_fielder["name"], f"{best_fielder['fielding']} fielding pts", f"({best_fielder['role']}, {best_fielder['owner']})"),
        ("Highest Average per Match", best_avg["name"], f"{best_avg['avg_per_match']} avg pts/match", f"({best_avg['role']}, {best_avg['owner']})"),
        ("Best Single Match Score", best_single["name"], f"{best_single['best_match_pts']} pts in one match", f"({best_single['role']}, {best_single['owner']})"),
        ("Most Matches Played", most_matches["name"], f"{most_matches['matches_played']} matches", f"({most_matches['role']}, {most_matches['owner']})"),
    ]

    # Role-wise MVPs
    for role in role_order:
        rp = [p for p in all_players if p["role"] == role]
        if rp:
            top = max(rp, key=lambda x: x["total"])
            records.append((
                f"Top {role}",
                top["name"],
                f"{top['total']} pts",
                f"({top['owner']})",
            ))

    _write_headers(ws7, row, ["Record", "Player", "Value", "Details"])
    row += 1
    for rec_title, rec_player, rec_value, rec_detail in records:
        ws7.cell(row=row, column=1, value=rec_title).border = border
        ws7.cell(row=row, column=1).alignment = left_a
        ws7.cell(row=row, column=1).font = record_font
        ws7.cell(row=row, column=2, value=rec_player).border = border
        ws7.cell(row=row, column=2).alignment = left_a
        ws7.cell(row=row, column=2).font = value_font
        ws7.cell(row=row, column=3, value=rec_value).border = border
        ws7.cell(row=row, column=3).alignment = left_a
        ws7.cell(row=row, column=4, value=rec_detail).border = border
        ws7.cell(row=row, column=4).alignment = left_a
        row += 1

    ws7.column_dimensions['A'].width = 30
    ws7.column_dimensions['B'].width = 26
    ws7.column_dimensions['C'].width = 26
    ws7.column_dimensions['D'].width = 24

    wb.save(filename)
    print(f"✓ Analytics exported to {filename}")
